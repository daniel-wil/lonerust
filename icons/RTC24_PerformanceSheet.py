#####################################################
# Author : Daniel Wilcox
# Create Date : June 1st 2023
# Revised : November 13th 2023
# Purpose : Automate RXC Performance Sheet
#####################################################

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service  
from selenium.webdriver.common.by import By
import threading
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date, datetime
from zoneinfo import ZoneInfo                          
import time
import pandas as pd
from google.oauth2 import service_account
from googleapiclient.discovery import build
import gspread
from gspread_formatting import *
from gspread import Cell
import openpyxl 
import matplotlib.colors as mcolors
import sys

def initialize_webdriver():
    # Code to initialize the Selenium webdriver
    # Return the webdriver instance
    options = Options()

    # Server specific arguments
    #options.add_argument("--remote-debugging-port=9222")   # each thread needs it's own port
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--headless=new")
    options.add_argument("--silent")
    options.add_argument('--log-level=3')
    options.add_argument("--disable-extensions")
    options.add_argument("--test-type")
    options.add_argument('--disable-dev-shm-usage')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])

    webdriver_path = ChromeDriverManager().install()
    service = Service(executable_path=webdriver_path)
    # initializing webdriver for Chrome with our options
    driver = webdriver.Chrome(service=service, options=options)
    return driver

def download_rsu_report(driver):
    # Code to scrape data using the webdriver
    # Return the scraped data

    # RSU Credentials
    rsu_username = "daniel.wilcox@runninglane.com"
    rsu_password = "vbe_unb9twf1KZA0tzq"
    rsu_report_url = "https://runsignup.com/Race/Participants/107646?csv=report"

    # getting RXC23 Reports webpage
    driver.get(rsu_report_url)
    driver.maximize_window()

    # find username/email field and send the username itself to the input field
    driver.find_element("name", "email").send_keys(rsu_username)

    # find password input field and insert password as well
    driver.find_element("name", "password").send_keys(rsu_password)

    # click login button
    driver.find_element(By.CSS_SELECTOR, "button.rsuBtn.rsuBtn--rsuThemed.fs-lg-2.submit").click()

    time.sleep(5/10)
    driver.close()


def download_directory(root_directory):
    # Code to return the RSU report download path

    todays_date = date.today()
    current_day = todays_date.day
    if (current_day < 10):
        current_day = "0" + str(current_day)
    else: current_day = str(todays_date.day)
    current_month = todays_date.month
    if (current_month < 10):
        current_month = "0" + str(current_month)
    else: current_month = str(current_month)
    current_year = str(todays_date.year)
    RTC24_Report_CSV = (root_directory + current_year + current_month +
                    current_day + "-RunningLaneTrackChampionships-participants.csv")
    RTC24_Report_XLSX = (root_directory + current_year + current_month +
                    current_day + "-RTC24Formatted.xlsx")
    return RTC24_Report_CSV, RTC24_Report_XLSX


def create_df(report_path_csv):
    # Read the CSV file into a pandas DataFrame
    df = pd.read_csv(report_path_csv)

    columns_to_drop = ['Middle Name', 'Street Address', 'City','Country', 'Zip Code']
    # Drop the specified columns from the DataFrame
    df = df.drop(columns=columns_to_drop)

    # Create column to store * for altitude adjusted
    altitude_converted_column = 'Altitude Adjusted*'
    df[altitude_converted_column] = None   
    return df

def multithread_altitude(num_threads, df, report_path_csv):
    # Calculate the number of rows each thread should process
        rows_per_thread = len(df) // num_threads
        # Creating and starting threads
        threads = []
        for i in range(num_threads):
            start_index = i * rows_per_thread
            # Ensure the last thread processes any remaining rows
            end_index = (i + 1) * rows_per_thread if i != num_threads - 1 else len(df)
            thread = threading.Thread(target=altitude_conversion, args=(report_path_csv, df, start_index, end_index))
            threads.append(thread)
            thread.start()

        # Wait for all threads to complete
        for thread in threads:
            thread.join()


def altitude_conversion(report_path_csv, df, start_index, end_index):
    # Code to scrape data using the webdriver

    # Specify columns to operate on
    eight_altitude_column = 'If your personal best time that you submitted for the 800m was run at altitude, please list the elevation of that town/ city. If your PR is run at sea level, leave this question blank.'
    mile_altitude_column = 'If your personal best time that you submitted for the mile was run at altitude, please list the elevation of that town/ city. If your PR is run at sea level, leave this question blank.'
    twomile_altitude_column = 'If your personal best time that you submitted for the 3200m was run at altitude, please list the elevation of that town/ city. If your PR is run at sea level, leave this question blank.'
   
    
    eight_pr_column = 'What is your personal best for the 800m?  Please give honest and verifiable seed times from the last calendar year. Indoor times do count.'
    mile_pr_column = 'What is your personal best for the mile. Please give honest and verifiable seed times from the last calendar year. Indoor times do count.'
    twomile_pr_column = 'What is your personal best for the 3200m. Please give honest and verifiable seed times from the last calendar year. Indoor times do count. '
    altitude_converted_column = 'Altitude Adjusted*'
    

    driver = initialize_webdriver()
    # getting FS Altitude Calculator 
    driver.get("https://www.finalsurge.com/altitude-conversion-calculator")
    driver.maximize_window()

    time.sleep(1/2)
    input_elements = driver.find_elements(By.CSS_SELECTOR, 'input.text-field.text-field--medium')
    # Check the operating system
    if sys.platform == 'darwin':  # darwin is the value for MacOS
        control_key = Keys.COMMAND
    else:
        control_key = Keys.CONTROL
    # Iterate over the values in the specified column
    # Print columns in df
    for index in range(start_index, end_index):
        # Decide which altitude value to send
        # If event column includes 800m, send 800m altitude
        if 'Championships 800m' in df.loc[index, 'Event']:
            altitude_column = eight_altitude_column
            pr_column = eight_pr_column
        elif 'Championships Mile' in df.loc[index, 'Event']:
            altitude_column = mile_altitude_column
            pr_column = mile_pr_column
        elif 'Championships 3200m' in df.loc[index, 'Event']:
            altitude_column = twomile_altitude_column
            pr_column = twomile_pr_column
        else: 
            continue
        altitude_value = df.loc[index, altitude_column]
        pr_time = df.loc[index, pr_column]
        # Turn altitude value into an int
        try:
            altitude_value = int(altitude_value)
        except ValueError:
            altitude_value = 0

        if(altitude_value >= 3000 and altitude_value != 0): 
            # Iterate through the input elements and send text based on position or other criteria
            for elements, input_element in enumerate(input_elements):
                if elements == 0:
                    input_element.send_keys(control_key + "a")  # Select all text in the input
                    input_element.send_keys(Keys.DELETE)         # Delete the selection
                    input_element.send_keys(altitude_value)
                elif elements == 1:
                    input_element.send_keys(control_key + "a")  # Select all text in the input
                    input_element.send_keys(Keys.DELETE)         # Delete the selection
                    # Decide what distance to send
                    # if event column includes 3200m, send 2.01168
                    if 'Championships 800m' in df.loc[index, 'Event']:
                        input_element.send_keys('0.497097')
                    elif 'Championships Mile' in df.loc[index, 'Event']:
                        input_element.send_keys('1')
                    elif 'Championships 3200m' in df.loc[index, 'Event']:
                        input_element.send_keys('1.988388')
                elif elements == 2:
                    input_element.send_keys(control_key + "a")  # Select all text in the input
                    input_element.send_keys(Keys.DELETE)         # Delete the selection
                    input_element.send_keys(pr_time)  
            time.sleep(3/10)
            converted_time = driver.find_element(By.CSS_SELECTOR, 'div.h3[data-v-38fbd95c]')
            converted_time = converted_time.text
            converted_time = ''.join(converted_time.split())[:-3].upper()

            # close browser after our manipulations
            df.loc[index, pr_column] = converted_time
            df.loc[index, altitude_converted_column] = '*'
            print(f"Index: {index} | Event: {df.loc[index, 'Event']} | Altitude: {altitude_value} | PR: {pr_time} | Converted Time: {converted_time}")
    # Save the DataFrame back to a CSV file
    driver.close()
    df.to_csv(report_path_csv, index=False)


def format_report(report_path_csv, report_path_xlsx):
    # Code to format the report
    df = pd.read_csv(report_path_csv)
    df.to_excel(report_path_xlsx, index=False)
    df = pd.read_excel(report_path_xlsx)


    altitudeMapping = {
        '800m Altitude': 'If your personal best time that you submitted for the 800m was run at altitude, please list the elevation of that town/ city. If your PR is run at sea level, leave this question blank.',
        'Mile Altitude': 'If your personal best time that you submitted for the mile was run at altitude, please list the elevation of that town/ city. If your PR is run at sea level, leave this question blank.',
        '3200m Altitude': 'If your personal best time that you submitted for the 3200m was run at altitude, please list the elevation of that town/ city. If your PR is run at sea level, leave this question blank.'
    }

    columns_to_drop = [altitudeMapping['800m Altitude'], altitudeMapping['Mile Altitude'], altitudeMapping['3200m Altitude']]
    df = df.drop(columns=columns_to_drop)

    # Create a mapping from original long names to shortened names
    eventsMap = {
        'RunningLane Track Championships 3200m Run (Girls)': '3200m Girls',
        'RunningLane Track Championships 3200m Run (Boys)': '3200m Boys',
        'RunningLane Track Championships Mile Run (Girls)': 'Mile Girls',
        'RunningLane Track Championships Mile Run (Boys)': 'Mile Boys',
        'RunningLane Track Championships 800m Run (Girls)': '800m Girls',
        'RunningLane Track Championships 800m Run (Boys)': '800m Boys',
        'RunningLane Track Championships 400m Run (Girls)': '400m Girls',
        'RunningLane Track Championships 400m Run (Boys)': '400m Boys',
        'RunningLane Track Championships 2000M Steeplechase Run (Girls)': 'Girls 2000m Steeplechase',
        'RunningLane Track Championships 2000M Steeplechase Run (Boys)': 'Boys 2000m Steeplechase',
        '3200m OPEN/YOUTH': '3200m Open',
        'Mile OPEN/YOUTH': 'Mile Open',
        '800m OPEN/YOUTH': '800m Open',
        '400m OPEN/YOUTH': '400m Open',
        '3000m Steeplechase OPEN/YOUTH': '3000m Steeplechase Open'
    }    

    # Map all column names to their shortened versions
    headerMapping = {
        'What grade are you in? (2023-2024 School Year)': 'Grade',
        'What is your personal best for the 400m. Please give honest and verifiable seed times from the last calendar year. Indoor times do count.': '400m',
        'What is your personal best for the 800m?  Please give honest and verifiable seed times from the last calendar year. Indoor times do count.': '800m',
        'What is your personal best for the mile. Please give honest and verifiable seed times from the last calendar year. Indoor times do count.': 'Mile',
        'What is your personal best for the 3200m. Please give honest and verifiable seed times from the last calendar year. Indoor times do count. ': '3200m'
    }


    # Rename the columns in the DataFrame
    df = df.rename(columns=headerMapping)

    # Create a dataframe for each event stored in EventsMap
    event_df = {event: df[df['Event'] == event] for event in eventsMap.keys()}
    # For each event df, sort the rows by their seed times for the event, if 400m sort by 400m times
    for event, df in event_df.items():
        # Sort the rows by the seed times for the event
        if '3200m' in event:
            # Enforce time duration format of hh:mm:ss for '3200m' column
            df['3200m'] = pd.to_datetime(df['3200m'], format='%H:%M:%S', errors='coerce').dt.time
            df = df.sort_values('3200m', ascending=True)
            columns_to_drop = ['Mile', '800m', '400m', 'Event']
            df = df.drop(columns=columns_to_drop)
            event_df[event] = df
        elif 'Mile' in event:
            # Enforce time duration format of hh:mm:ss for 'Mile' column
            df['Mile'] = pd.to_datetime(df['Mile'], format='%H:%M:%S', errors='coerce').dt.time
            df = df.sort_values('Mile', ascending=True)
            columns_to_drop = ['3200m', '800m', '400m', 'Event']
            df = df.drop(columns=columns_to_drop)
            event_df[event] = df
        elif '800m' in event:
            # Enforce time duration format of hh:mm:ss for '800m' column
            df['800m'] = pd.to_datetime(df['800m'], format='%H:%M:%S', errors='coerce').dt.time
            df = df.sort_values('800m', ascending=True)
            columns_to_drop = ['3200m', 'Mile', '400m', 'Event']
            df = df.drop(columns=columns_to_drop) 
            event_df[event] = df  
        elif '400m' in event:
            # Enforce time duration format of hh:mm:ss for '400m' column
            df['400m'] = pd.to_datetime(df['400m'], format='%H:%M:%S', errors='coerce').dt.time
            df = df.sort_values('400m', ascending=True)
            columns_to_drop = ['3200m', 'Mile', '800m', 'Event']
            df = df.drop(columns=columns_to_drop)
            event_df[event] = df
        
    # Store each dataframe in a separate sheet in the Excel file using the mapped sheet names
    with pd.ExcelWriter(report_path_xlsx) as writer:
        for event, df in event_df.items():
            df.to_excel(writer, sheet_name=eventsMap[event], index=False)
        

def upload_to_google_sheets(report_path_xlsx, credentials_path_json):
    # Code to upload the Excel file to Google Sheets
    # Specify the name of your Google Sheets spreadsheet
    spreadsheet_name = 'RTC 2024 Performance List'
    print(f"Uploading to Google Sheets: {spreadsheet_name}")
    # Authenticate using the service account credentials
    creds = service_account.Credentials.from_service_account_file(
        credentials_path_json,
        scopes=['https://www.googleapis.com/auth/drive','https://www.googleapis.com/auth/spreadsheets']
    )
    client = gspread.authorize(creds)
    print("Authenticated with Google Sheets")
    spreadsheet = client.open(spreadsheet_name)
    # Load the XLSX file
    workbook = openpyxl.load_workbook(filename=report_path_xlsx)
    # Iterate over each sheet in the workbook
    # Loop through each sheet in the XLSX file
    for sheet_name in workbook.sheetnames:
        # Access the current sheet in the XLSX file
        print(f"Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]

        # Convert rows and columns into lists to get their lengths
        rows = list(sheet.iter_rows(values_only=True))

        # Check if the sheet exists in Google Sheets
        try:
            worksheet = spreadsheet.worksheet(sheet_name)
            print(f"Found existing worksheet: {sheet_name}")
        except gspread.exceptions.WorksheetNotFound:
            # If the sheet doesn't exist, create it
            worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=len(rows), cols=len(rows[0]))
            print(f"Created new worksheet: {sheet_name}")

        # Clear the contents of the Google Sheets worksheet
        worksheet.clear()

        # Batch size for appending rows
        batch_size = 1000  # Adjust as needed

        # Loop through rows in the current XLSX sheet and append them in batches
        for start_index in range(0, len(rows), batch_size):
            batch = rows[start_index:start_index + batch_size]
            worksheet.append_rows(batch, value_input_option='USER_ENTERED')
            # Pad columns and rows to ensure the worksheet is large enough for viewing
            worksheet.resize(rows=len(rows) + 1000, cols=len(rows[0]) + 50)
            # Bold the header row
            # Create a CellFormat object with bold text
            bold_format = CellFormat(textFormat=TextFormat(bold=True))

            # Bold the header row
            format_cell_range(worksheet, 'A1:Z1', bold_format)


def main():
    driver = initialize_webdriver()
    download_rsu_report(driver)
    # MacOS: "/Users/danielwilcox/Downloads/", Windows:
    report_path_csv, report_path_xlsx = download_directory("/Users/danielwilcox/Downloads/")
    df = create_df(report_path_csv)
    multithread_altitude(1, df, report_path_csv)
    format_report(report_path_csv, report_path_xlsx)
    # '/home/danielwilcox/credentials.json' 'C:\\Users\\Daniel\\Documents\\VSCODE\\RXC23\\Performance Scraping\\credentials.json' 
    # '/home/ec2-user/credentials.json'
    upload_to_google_sheets(report_path_xlsx, '/Users/danielwilcox/Documents/vscode/RunningLane/RTC24/credentials.json')


if __name__ == "__main__":
    main()